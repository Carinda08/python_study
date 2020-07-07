import openpyxl

# open test.xlsx
workbook = openpyxl.load_workbook('test.xlsx')

# get sheet list
sheets = workbook.get_sheet_names()
for st in sheets:
    print(st)

sheet = workbook.get_sheet_by_name('Fruit')
# get sheet size
row = sheet.max_row
col = sheet.max_column
print("size : %s %s " % (row, col))

# get cell
# use name
a1 = sheet['A1']
print(a1.value)
b1 = sheet['B1']
print(b1.value)

# use index
for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)

# use row or column
for cell in list(sheet.columns)[0]:
    print(cell.value)

for cell in sheet['1']:
    print(cell.value)
