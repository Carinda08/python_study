import openpyxl, pprint

db = {}

def anaxls(xlsfile):
    print("Starting to analyze " + xlsfile )
    # open test.xlsx
    workbook = openpyxl.load_workbook(xlsfile)
    shtNames = workbook.get_sheet_names()

    for na in shtNames:
        tmpSheet = workbook.get_sheet_by_name(na)
        anasht(tmpSheet)
    print(xlsfile + " Done!")

def anasht(xlsSheet):
    print("Staring to analyze " + xlsSheet.title)
    for i in range(2, xlsSheet.max_row+1):
        state = xlsSheet.cell(i, 2).value
        country = xlsSheet.cell(i, 3).value
        popNbr = xlsSheet.cell(i, 4).value

        db.setdefault(state, {})
        db[state].setdefault(country, {'count': 0, 'nbr': 0})
        db[state][country]['count'] += 1
        db[state][country]['nbr']+= int(popNbr)
    print(xlsSheet.title + " Done!")

def outputDB(outfile):
    print("Output db data to file " + outfile)
    with open(outfile, 'w') as fo:
        fo.write('db = ' + pprint.pformat(db))
        fo.close()

def outputxls(outfile):
    print("Output db data to excel " + outfile)



if __name__ == '__main__':
    anaxls('censuspopdata.xlsx')
    print(db)
    outputDB('census.py')
